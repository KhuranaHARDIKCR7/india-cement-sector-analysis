import os
import glob
import shutil
import pandas as pd
import numpy as np
import openpyxl

base_dir = os.path.dirname(os.path.abspath(__file__))
out_dir_base = os.path.join(base_dir, "data", "processed")

years_to_process = [
    ('2009_10', '200910')
]

a_rename_maps = {
    '2009_10': {
        'YR': 'Year', 'A_Itm2': 'PSL', 'A_Itm3': 'Scheme', 'A_Itm4': 'INC4digit',
        'A_Itm5': 'INC5digit', 'A_Itm7': 'State', 'A_Itm8': 'District', 'A_Itm9': 'Rural_Urban',
        'A_Itm10': 'RO_SRO', 'A_Itm11': 'Unit', 'A_Itm12': 'Status_Unit', 'E_Itm10': 'Bonus',
        'E_Itm11': 'ProvidentFund', 'E_Itm12': 'Welfare', 'E_Itm13a': 'MWorkingdays',
        'E_Itm13b': 'NMWorkingdays', 'E_Itm13c': 'TWorkingdays', 'E_Itm14': 'CostofProd',
        'J_Itm13': 'Share', 'WGT': 'Multilplier'
    },
    'default': {
        'NIC4digit': 'INC4digit', 'NIC5digit': 'INC5digit', 'StateCode': 'State',
        'NoofUnits': 'Unit', 'Statusofunit': 'Status_Unit'
    }
}

block_mappings = {
    'h': {
        '2009_10': {
            'YR': 'Year', 'BLK': 'BLK', 'DSL': 'DSL', 'H_Itm1': 'Sno',
            'H_Itm3': 'ItemCode', 'H_Itm4': 'Unitcode', 'H_Itm5': 'QtyCons', 'H_Itm6': 'PurVal'
        },
        'default': {
            'DSL': 'DSL', 'Sno': 'Sno', 'ItemCode': 'ItemCode', 'Unitcode': 'Unitcode',
            'QtyCons': 'QtyCons', 'PurVal': 'PurVal'
        }
    },
    'i': {
        '2009_10': {
            'YR': 'Year', 'BLK': 'BLK', 'DSL': 'DSL', 'I_Itm1': 'Sno',
            'I_Itm3': 'ItemCode', 'I_Itm4': 'Unitcode', 'I_Itm5': 'QtyCons', 'I_Itm6': 'Purvaldel',
            'I_Itm7': 'Rateperunit'
        },
        'default': {
            'DSL': 'DSL', 'Sno': 'Sno', 'ItemCode': 'ItemCode', 'Unitcode': 'Unitcode',
            'QtyCons': 'QtyCons', 'Purvaldel': 'Purvaldel', 'Rateperunit': 'Rateperunit'
        }
    },
    'j': {
        '2009_10': {
            'YR': 'Year', 'BLK': 'BLK', 'DSL': 'DSL', 'J_Itm1': 'Sno',
            'J_Itm3': 'ItemCode', 'J_Itm4': 'Unitcode', 'J_Itm5': 'QtyCons', 'J_Itm6': 'QtySold',
            'J_Itm7': 'Grosssalval', 'J_Itm8': 'ExciseDuty', 'J_Itm9': 'SalesTax', 'J_Itm10': 'Others',
            'J_Itm11': 'Total', 'J_Itm12': 'NetSaleval', 'J_Itm13': 'ExfactvalOutput'
        },
        'default': {
            'DSL': 'DSL', 'Sno': 'Sno', 'ItemCode': 'ItemCode', 'Unitcode': 'Unitcode',
            'QtyManuf': 'QtyCons', 'QtySold': 'QtySold', 'Grosssalval': 'Grosssalval',
            'ExciseDuty': 'ExciseDuty', 'SalesTax': 'SalesTax', 'Others': 'Others',
            'Total': 'Total', 'NetSaleval': 'NetSaleval', 'ExfactvalOutput': 'ExfactvalOutput'
        }
    }
}

def prepare_blka(year_folder, year_sfx):
    folder_path = os.path.join(base_dir, f"ASI_DATA_{year_folder}_CSV")
    out_folder = os.path.join(out_dir_base, f"ASI_DATA_{year_folder}_CSV")
    os.makedirs(out_folder, exist_ok=True)
    
    if year_folder == '2009_10':
        blka_files = glob.glob(os.path.join(folder_path, "A-IDENTIFICATION*.csv"))
    else:
        blka_files = glob.glob(os.path.join(folder_path, "blka*.csv"))
        if not blka_files:
            blka_files = glob.glob(os.path.join(folder_path, "blkA*.csv"))
            
    df_a = pd.read_csv(blka_files[0], encoding='utf-8', low_memory=False)
    
    if year_folder == '2009_10':
        df_a.rename(columns=a_rename_maps['2009_10'], inplace=True)
    else:
        df_a.rename(columns=a_rename_maps['default'], inplace=True)
        
    dsl_col = [c for c in df_a.columns if 'DSL' in c.upper()][0]
    df_a[dsl_col] = pd.to_numeric(df_a[dsl_col], errors='coerce')
    
    out_blka = os.path.join(out_folder, f"blka{year_sfx}.xlsx")
    with pd.ExcelWriter(out_blka, engine='openpyxl') as writer:
        df_a.to_excel(writer, sheet_name=f'blka{year_sfx}', index=False)
    print(f"Prepared Block A for {year_folder} at {out_blka}")

def replicate_block_openpyxl(year_folder, year_sfx, block_char):
    block_lower = block_char.lower()
    folder_path = os.path.join(base_dir, "data", "raw", f"ASI_DATA_{year_folder}_CSV")
    out_folder = os.path.join(out_dir_base, f"ASI_DATA_{year_folder}_CSV")
    
    # Copy template
    ref_name = f"blk{block_lower}201415.xlsx"
    if block_lower == 'h': ref_name = 'blkH201415.xlsx'
    elif block_lower == 'i': ref_name = 'blkI201415.xlsx'
    
    ref_path = os.path.join(out_dir_base, "ASI_DATA_2014_15_CSV", ref_name)
    dest_path = os.path.join(out_folder, f"blk{block_lower}{year_sfx}.xlsx")
    
    shutil.copy(ref_path, dest_path)
    print(f"Copied template to {dest_path}")
    
    # 1. Load CSV file
    if year_folder == '2009_10':
        prefix = {'h': 'H-INPUT', 'i': 'I-INPUT', 'j': 'J-PRODUCTS'}[block_lower]
        csv_files = glob.glob(os.path.join(folder_path, f"{prefix}*.csv"))
    else:
        csv_files = glob.glob(os.path.join(folder_path, f"blk{block_char}*.csv"))
        if not csv_files:
            csv_files = glob.glob(os.path.join(folder_path, f"blk{block_char.upper()}*.csv"))
            
    if not csv_files:
        print(f"No CSV found for {year_folder} Block {block_char}")
        return
        
    df = pd.read_csv(csv_files[0], encoding='utf-8', low_memory=False)
    df.rename(columns=block_mappings[block_lower]['2009_10' if year_folder == '2009_10' else 'default'], inplace=True)
    
    # Load reference list from template to perform sorting (only for H and I)
    if block_lower in ['h', 'i']:
        wb_ref = openpyxl.load_workbook(ref_path, read_only=True)
        ws_ref = wb_ref.active
        
        ref_npcms = []
        ref_col = 18 if block_lower == 'h' else 17
        for r in range(2, ws_ref.max_row + 1):
            v = ws_ref.cell(r, ref_col).value
            if v is None: break
            ref_npcms.append(int(v))
        wb_ref.close()
        
        # Fetch Block A to calculate NIC/MULT for sorting
        blka_file = os.path.join(out_folder, f"blka{year_sfx}.xlsx")
        df_a = pd.read_excel(blka_file)
        
        # Standardize types for merging
        df['DSL'] = pd.to_numeric(df['DSL'], errors='coerce')
        df_a['DSL'] = pd.to_numeric(df_a['DSL'], errors='coerce')
        
        # Merge with Block A to get NIC5digit
        df_merged = pd.merge(df, df_a[['DSL', 'INC5digit']], on='DSL', how='left')
    else:
        df_merged = df.copy()

    
    # Sorting logic
    if block_lower in ['h', 'i']:
        df_merged['ItemCode_num'] = pd.to_numeric(df_merged['ItemCode'], errors='coerce')
        df_merged['NIC_num'] = pd.to_numeric(df_merged['INC5digit'], errors='coerce')
        
        if year_folder == '2009_10':
            ref_npcms_strs = {str(c)[:5] for c in ref_npcms}
            def is_item_target(item):
                item_str = str(item).strip()
                if '.' in item_str:
                    item_str = item_str.split('.')[0]
                return item_str in ref_npcms_strs or (item_str.isdigit() and str(int(item_str)) in ref_npcms_strs)
            
            if block_lower == 'h':
                df_merged['is_target'] = df_merged.apply(
                    lambda r: (23940 <= r['NIC_num'] <= 23950) and is_item_target(r['ItemCode']), axis=1
                )
            elif block_lower == 'i':
                df_merged['is_target'] = df_merged.apply(
                    lambda r: is_item_target(r['ItemCode']), axis=1
                )
        else:
            if block_lower == 'h':
                df_merged['is_target'] = df_merged.apply(
                    lambda r: (23940 <= r['NIC_num'] <= 23950) and (r['ItemCode_num'] in ref_npcms), axis=1
                )
            elif block_lower == 'i':
                df_merged['is_target'] = df_merged['ItemCode_num'].isin(ref_npcms)
            
        df_merged.sort_values(by='is_target', ascending=False, kind='mergesort', inplace=True)
        num_target_rows = df_merged['is_target'].sum()
        print(f"Block {block_char.upper()}: sorted {num_target_rows} target rows to top.")
            
        df = df_merged.drop(columns=['INC5digit', 'ItemCode_num', 'NIC_num', 'is_target'], errors='ignore')
    else:
        num_target_rows = len(df_merged)
        df = df_merged.copy()

        
    # Load destination file with openpyxl
    wb = openpyxl.load_workbook(dest_path)
    sheet_ref_name = f"blk{block_lower}201415"
    if block_lower == 'h': sheet_ref_name = 'blkh201415'
    elif block_lower == 'i': sheet_ref_name = 'blkI201415'
    
    ws = wb[sheet_ref_name]
    ws.title = f"blk{block_lower}{year_sfx}"
    
    old_last_row = ws.max_row
    new_last_row = len(df) + 1
    
    # We define the column mappings (1-indexed for openpyxl)
    if block_lower == 'h':
        raw_cols = {
            1: df['Year'].values,
            2: df['BLK'].values,
            3: df['DSL'].values,
            6: df['Sno'].values,
            7: df['ItemCode'].values,
            9: df['Unitcode'].values,
            12: df['QtyCons'].values,
            14: df['PurVal'].values
        }
        formulas = {
            4: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$H,6,FALSE)",
            5: lambda r: f"=VLOOKUP(D{r},Sheet1!$A:$B,2,FALSE)",
            8: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$G,5,FALSE)",
            10: lambda r: f"=VLOOKUP(I{r},Sheet1!D:E,2,FALSE)",
            11: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$V,20,FALSE)",
            13: lambda r: f"=K{r}*L{r}",
            15: lambda r: f"=K{r}*N{r}" # Est Value in Block H column O
        }
    elif block_lower == 'i':
        raw_cols = {
            1: df['Year'].values,
            2: df['BLK'].values,
            3: df['DSL'].values,
            4: df['Sno'].values,
            7: df['ItemCode'].values,
            9: df['Unitcode'].values,
            11: df['QtyCons'].values,
            13: df['Purvaldel'].values,
            15: df['Rateperunit'].values
        }
        formulas = {
            5: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$H,6,FALSE)",
            6: lambda r: f"=VLOOKUP(E{r},Sheet1!A:B,2,FALSE)",
            8: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$G,5,FALSE)",
            10: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$V,20,FALSE)",
            12: lambda r: f"=J{r}*K{r}",
            14: lambda r: f"=J{r}*M{r}"
        }
    elif block_lower == 'j':
        raw_cols = {
            1: df['Year'].values,
            2: df['BLK'].values,
            3: df['DSL'].values,
            4: df['Sno'].values,
            7: df['ItemCode'].values,
            9: df['Unitcode'].values,
            12: df['QtyCons'].values,
            14: df['QtySold'].values,
            15: df['Grosssalval'].values,
            16: df['ExciseDuty'].values,
            17: df['SalesTax'].values,
            18: df['Others'].values,
            19: df['Total'].values,
            20: df['NetSaleval'].values,
            21: df['ExfactvalOutput'].values
        }
        formulas = {
            5: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$H,6,FALSE)",
            6: lambda r: f"=VLOOKUP(E{r},Sheet1!$A:$B,2,FALSE)",
            8: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$G,5,FALSE)",
            10: lambda r: f"=VLOOKUP(I{r},Sheet1!D:E,2,FALSE)",
            11: lambda r: f"=VLOOKUP(C{r},'[blka{year_sfx}.xlsx]blka{year_sfx}'!$C:$V,20,FALSE)",
            13: lambda r: f"=K{r}*L{r}"
        }

    print("Writing cell values and formulas...")
    for col_idx, vals in raw_cols.items():
        for i, val in enumerate(vals):
            r = i + 2
            cell_val = '' if pd.isna(val) or val is None else val
            ws.cell(row=r, column=col_idx).value = cell_val

    # Write formulas
    for col_idx, formula_func in formulas.items():
        for r in range(2, new_last_row + 1):
            ws.cell(row=r, column=col_idx).value = formula_func(r)

    # Delete extra rows
    if old_last_row > new_last_row:
        print(f"Deleting extra rows: {new_last_row + 1} to {old_last_row}")
        ws.delete_rows(new_last_row + 1, old_last_row - new_last_row)
        
    # Update Pivot Table Source range and refresh on load
    if ws._pivots:
        for pivot in ws._pivots:
            cache = pivot.cache
            if cache.cacheSource and cache.cacheSource.worksheetSource:
                ws_src = cache.cacheSource.worksheetSource
                ws_src.sheet = f"blk{block_lower}{year_sfx}"
                
                # Pivot table source range now covers exactly the target partition rows!
                if block_lower == 'h':
                    ws_src.ref = f"A1:P{num_target_rows + 1}"
                elif block_lower == 'i':
                    ws_src.ref = f"A1:O{num_target_rows + 1}"
                elif block_lower == 'j':
                    ws_src.ref = f"B1:U{new_last_row}"
                    
                cache.refreshOnLoad = True
                print(f"Updated Pivot Table Source to {ws_src.sheet}!{ws_src.ref}")
                
    wb.save(dest_path)
    print(f"Saved {dest_path}")

if __name__ == "__main__":
    for y_folder, y_sfx in years_to_process:
        print(f"=== Processing Year {y_folder} ===")
        for b in ['j']:
            replicate_block_openpyxl(y_folder, y_sfx, b)
